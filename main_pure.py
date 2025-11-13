# -*- coding: utf-8 -*-
"""
外箱标签打印系统 - 纯数据写入版 v2.2
功能：只写入数据，不修改模板样式
GitHub: https://github.com/OuterBoxPrinter/PureData
"""

import sys
import os
import sqlite3
import hashlib
import json
from datetime import datetime
import win32com.client
import openpyxl
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTableWidget, QTableWidgetItem, QLabel, QLineEdit,
    QComboBox, QDateEdit, QTextEdit, QMessageBox, QFileDialog,
    QTabWidget, QInputDialog, QHeaderView, QGroupBox, QFormLayout
)
from PyQt5.QtCore import QDate, Qt
from PyQt5.QtGui import QFont

# ==================== 配置加载 ====================
CONFIG_FILE = "config_pure.json"
DEFAULT_CONFIG = {
    "mode": "pure_data",
    "write_only": True,
    "modify_style": False,
    "save_template": False,
    "field_mapping": {
        "mingcheng": "name",
        "guige": "spec",
        "xinghao": "model",
        "yanse": "color",
        "SKU": "sku",
        "69": "code69",
        "shuliang": "quantity",
        "zhongliang": "weight"
    },
    "template_path": "example_template.btw"
}

if not os.path.exists(CONFIG_FILE):
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(DEFAULT_CONFIG, f, ensure_ascii=False, indent=2)

with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
    config = json.load(f)

# ==================== 数据库初始化 ====================
DB_FILE = "outerbox.db"

def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS products (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT, spec TEXT, model TEXT, color TEXT,
        sku TEXT, code69 TEXT, quantity INTEGER, weight REAL,
        template TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS print_records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        product_id INTEGER,
        sn TEXT UNIQUE,
        box_no TEXT,
        print_date TEXT,
        print_time TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT
    )''')
    c.execute("INSERT OR IGNORE INTO settings VALUES ('box_rule', 'BOX-{YYYY}{MM}{DD}-{SEQ:0000}')")
    c.execute("INSERT OR IGNORE INTO settings VALUES ('box_seq', '0')")
    conn.commit()
    conn.close()

# ==================== 工具函数 ====================
def get_setting(key, default=""):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("SELECT value FROM settings WHERE key=?", (key,))
    row = c.fetchone()
    conn.close()
    return row[0] if row else default

def set_setting(key, value):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO settings VALUES (?, ?)", (key, value))
    conn.commit()
    conn.close()

def is_sn_printed(sn):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("SELECT 1 FROM print_records WHERE sn=?", (sn,))
    exists = c.fetchone() is not None
    conn.close()
    return exists

# ==================== 主窗口 ====================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("外箱标签打印系统 - 纯数据写入版 v2.2")
        self.setGeometry(100, 100, 1100, 700)
        self.setStyleSheet("font-family: Microsoft YaHei; background:#f8f9fa;")
        init_db()
        self.init_ui()

    def init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QHBoxLayout(central)
        layout.setContentsMargins(0, 0, 0, 0)

        # 左侧菜单
        menu = QWidget()
        menu.setFixedWidth(180)
        menu.setStyleSheet("background:#2c3e50; border-radius:8px;")
        menu_layout = QVBoxLayout(menu)
        buttons = [
            ("产品管理", self.show_product_tab),
            ("打印标签", self.show_print_tab),
            ("打印记录", self.show_record_tab),
            ("设置", self.show_setting_tab),
        ]
        for text, func in buttons:
            btn = QPushButton(text)
            btn.setStyleSheet("""
                QPushButton {background:#34495e; color:white; border:none; padding:15px; text-align:left; font-size:14px;}
                QPushButton:hover {background:#1abc9c;}
            """)
            btn.clicked.connect(func)
            menu_layout.addWidget(btn)
        menu_layout.addStretch()
        layout.addWidget(menu)

        # 右侧内容
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs, 1)

        self.product_page = ProductPage()
        self.print_page = PrintPage()
        self.record_page = RecordPage()
        self.setting_page = SettingPage()

        self.show_product_tab()

    def show_product_tab(self): self.tabs.clear(); self.tabs.addTab(self.product_page, "产品管理")
    def show_print_tab(self): self.tabs.clear(); self.tabs.addTab(self.print_page, "打印标签")
    def show_record_tab(self): self.tabs.clear(); self.tabs.addTab(self.record_page, "打印记录")
    def show_setting_tab(self): self.tabs.clear(); self.tabs.addTab(self.setting_page, "设置")

# ==================== 产品管理页面 ====================
class ProductPage(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        top = QHBoxLayout()
        top.addWidget(QPushButton("导入Excel", clicked=self.import_excel))
        top.addWidget(QPushButton("导出Excel", clicked=self.export_excel))
        top.addStretch()
        layout.addLayout(top)

        self.table = QTableWidget(0, 9)
        self.table.setHorizontalHeaderLabels(["名称", "规格", "型号", "颜色", "SKU", "69码", "数量", "重量", "模板"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.table)
        self.load_products()

    def load_products(self):
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT name,spec,model,color,sku,code69,quantity,weight,template FROM products")
        rows = c.fetchall()
        self.table.setRowCount(len(rows))
        for i, row in enumerate(rows):
            for j, val in enumerate(row):
                self.table.setItem(i, j, QTableWidgetItem(str(val or "")))
        conn.close()

    def import_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择Excel", "", "Excel (*.xlsx)")
        if not path: return
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # 名称不为空
                c.execute("""INSERT OR REPLACE INTO products 
                    (name,spec,model,color,sku,code69,quantity,weight,template)
                    VALUES (?,?,?,?,?,?,?,?,?)""", row)
        conn.commit()
        conn.close()
        self.load_products()
        QMessageBox.information(self, "成功", "导入完成")

    def export_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "保存Excel", "products.xlsx", "Excel (*.xlsx)")
        if not path: return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["名称", "规格", "型号", "颜色", "SKU", "69码", "数量", "重量", "模板"])
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT name,spec,model,color,sku,code69,quantity,weight,template FROM products")
        for row in c.fetchall():
            ws.append(row)
        wb.save(path)
        conn.close()
        QMessageBox.information(self, "成功", "导出完成")

# ==================== 打印标签页面 ====================
class PrintPage(QWidget):
    def __init__(self):
        super().__init__()
        self.selected_product = None
        self.sn_list = []
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)

        # 产品选择
        top = QHBoxLayout()
        self.search = QLineEdit()
        self.search.setPlaceholderText("输入69码搜索")
        self.search.textChanged.connect(self.search_product)
        top.addWidget(QLabel("搜索:"))
        top.addWidget(self.search)
        self.product_list = QComboOwners()
        top.addWidget(self.product_list, 1)
        layout.addLayout(top)

        # 产品信息
        self.info = QLabel("请选择产品")
        self.info.setStyleSheet("background:#e9ecef; padding:10px; border-radius:5px;")
        layout.addWidget(self.info)

        # 生产日期
        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("生产日期:"))
        self.date_edit = QDateEdit()
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)
        date_layout.addWidget(self.date_edit)
        date_layout.addStretch()
        layout.addLayout(date_layout)

        # SN 输入区
        sn_group = QGroupBox("SN 录入（扫描枪回车提交）")
        sn_layout = QVBoxLayout(sn_group)
        self.sn_input = QTextEdit()
        self.sn_input.setPlaceholderText("每行一个SN，达到整箱数量自动打印")
        self.sn_input.setFocus()
        sn_layout.addWidget(self.sn_input)
        btn_layout = QHBoxLayout()
        btn_layout.addWidget(QPushButton("清除未打印", clicked=self.clear_unprinted))
        btn_layout.addWidget(QPushButton("手动打印", clicked=self.manual_print))
        sn_layout.addLayout(btn_layout)
        layout.addWidget(sn_group)

        self.sn_input.textChanged.connect(self.check_sn_count)
        self.product_list.currentIndexChanged.connect(self.load_product)
        self.load_product_list()

    def load_product_list(self):
        self.product_list.clear()
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT id, name, code69 FROM products")
        for pid, name, code in c.fetchall():
            self.product_list.addItem(f"{code} - {name}", pid)
        conn.close()

    def search_product(self, text):
        self.product_list.clear()
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT id, name, code69 FROM products WHERE code69 LIKE ?", (f"%{text}%",))
        for pid, name, code in c.fetchall():
            self.product_list.addItem(f"{code} - {name}", pid)
        conn.close()

    def load_product(self):
        pid = self.product_list.currentData()
        if not pid: return
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT * FROM products WHERE id=?", (pid,))
        row = c.fetchone()
        conn.close()
        if row:
            self.selected_product = row
            self.info.setText(f"产品: {row[1]} | 规格: {row[2]} | 整箱数量: {row[7]}")
            self.check_sn_count()

    def check_sn_count(self):
        if not self.selected_product: return
        sns = [line.strip() for line in self.sn_input.toPlainText().split('\n') if line.strip()]
        count = len(sns)
        target = self.selected_product[7]  # quantity
        if count == target:
            self.print_label(sns)

    def manual_print(self):
        if not self.selected_product:
            QMessageBox.warning(self, "错误", "请先选择产品")
            return
        sns = [line.strip() for line in self.sn_input.toPlainText().split('\n') if line.strip()]
        if len(sns) != self.selected_product[7]:
            QMessageBox.warning(self, "错误", f"SN数量必须为 {self.selected_product[7]}")
            return
        self.print_label(sns)

    def clear_unprinted(self):
        self.sn_input.clear()

    def print_label(self, sns):
        if any(is_sn_printed(sn) for sn in sns):
            QMessageBox.warning(self, "错误", "有SN已打印过！")
            return

        try:
            bt = win32com.client.Dispatch("BarTender.Application")
            bt.Visible = False
            doc = bt.Documents.Open(self.selected_product[9])  # template

            # === 只写数据，不改样式 ===
            mapping = config["field_mapping"]
            product_data = {
                "name": self.selected_product[1],
                "spec": self.selected_product[2],
                "model": self.selected_product[3],
                "color": self.selected_product[4],
                "sku": self.selected_product[5],
                "code69": self.selected_product[6],
                "quantity": str(self.selected_product[7]),
                "weight": str(self.selected_product[8])
            }
            for var_name, key in mapping.items():
                try:
                    doc.Variables(var_name).Value = product_data.get(key, "")
                except: pass

            # === 写入 SN ===
            for i, sn in enumerate(sns):
                try:
                    doc.Variables(str(i + 1)).Value = sn
                except: break

            doc.PrintOut(False, False)
            doc.Close(2)  # 不保存
            bt.Quit()

            # === 存档 ===
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            box_no = self.generate_box_no()
            for sn in sns:
                c.execute("INSERT INTO print_records VALUES (NULL,?,?,?,?,?)",
                          (self.selected_product[0], sn, box_no,
                           self.date_edit.date().toString("yyyy-MM-dd"),
                           datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            conn.commit()
            conn.close()

            QMessageBox.information(self, "成功", f"打印完成，箱号: {box_no}")
            self.sn_input.clear()
        except Exception as e:
            QMessageBox.critical(self, "打印失败", str(e))

    def generate_box_no(self):
        rule = get_setting("box_rule", "BOX-{YYYY}{MM}{DD}-{SEQ:0000}")
        now = datetime.now()
        seq = int(get_setting("box_seq", "0")) + 1
        set_setting("box_seq", str(seq))
        rule = rule.replace("{YYYY}", now.strftime("%Y"))
        rule = rule.replace("{MM}", now.strftime("%m"))
        rule = rule.replace("{DD}", now.strftime("%d"))
        rule = rule.replace("{SEQ:0000}", f"{seq:04d}")
        return rule

# ==================== 打印记录页面 ====================
class RecordPage(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        search = QHBoxLayout()
        search.addWidget(QLabel("SN:"))
        self.sn_search = QLineEdit()
        search.addWidget(self.sn_search)
        search.addWidget(QLabel("69码:"))
        self.code_search = QLineEdit()
        search.addWidget(self.code_search)
        search.addWidget(QPushButton("查询", clicked=self.query))
        search.addWidget(QPushButton("导出Excel", clicked=self.export_excel))
        layout.addLayout(search)

        self.table = QTableWidget(0, 8)
        self.table.setHorizontalHeaderLabels(["SN", "箱号", "生产日期", "打印时间", "名称", "规格", "69码", "SKU"])
        layout.addWidget(self.table)
        self.query()

    def query(self):
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        sql = """SELECT r.sn, r.box_no, r.print_date, r.print_time,
                        p.name, p.spec, p.code69, p.sku
                 FROM print_records r JOIN products p ON r.product_id = p.id WHERE 1=1"""
        params = []
        if self.sn_search.text():
            sql += " AND r.sn LIKE ?"
            params.append(f"%{self.sn_search.text()}%")
        if self.code_search.text():
            sql += " AND p.code69 LIKE ?"
            params.append(f"%{self.code_search.text()}%")
        c.execute(sql, params)
        rows = c.fetchall()
        self.table.setRowCount(len(rows))
        for i, row in enumerate(rows):
            for j, val in enumerate(row):
                self.table.setItem(i, j, QTableWidgetItem(str(val)))
        conn.close()

    def export_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "导出", "records.xlsx", "Excel (*.xlsx)")
        if not path: return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["SN", "箱号", "生产日期", "打印时间", "名称", "规格", "69码", "SKU"])
        for i in range(self.table.rowCount()):
            row = [self.table.item(i, j).text() for j in range(8)]
            ws.append(row)
        wb.save(path)
        QMessageBox.information(self, "成功", "导出完成")

# ==================== 设置页面 ====================
class SettingPage(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        btn = QPushButton("进入设置（密码：admin123）")
        btn.clicked.connect(self.check_password)
        layout.addWidget(btn)
        layout.addStretch()

    def check_password(self):
        pwd, ok = QInputDialog.getText(self, "密码", "请输入密码:", QLineEdit.Password)
        if ok and pwd == "admin123":
            self.open_settings()
        else:
            QMessageBox.warning(self, "错误", "密码错误")

    def open_settings(self):
        win = QWidget()
        win.setWindowTitle("系统设置")
        win.setGeometry(300, 300, 500, 300)
        layout = QVBoxLayout(win)
        layout.addWidget(QLabel("功能开发中..."))
        win.show()

# ==================== 启动程序 ====================
if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())